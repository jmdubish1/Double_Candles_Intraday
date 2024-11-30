import numpy as np
import pandas as pd
import openpyxl
import os
from datetime import datetime, timedelta
from tensorflow.keras.metrics import Precision, Recall, AUC
from openpyxl.drawing.image import Image
import neural_network.nn_tools.loss_functions as lf
from neural_network.nn_tools.model_tools2 import TemperatureScalingLayer
import keras
import pickle


class SaveHandler:
    def __init__(self, mkt_data, lstm_model, data_params):
        self.mkt_data = mkt_data
        self.lstm_model = lstm_model
        self.process_handler = None
        self.data_params = data_params
        self.save_file = None
        self.trade_metrics = None

        self.test_date = None
        self.end_date = None
        self.train_date = None
        
        self.model_summary = None

        self.param_folder = ''
        self.data_folder = ''
        self.model_folder = ''
        self.main_train_path = ''
        self.previous_model_path = None
        self.model_save_path = None

    def check_create_model_folder(self):
        self.param_folder = \
            (f'{self.process_handler.data_params.trade_dat_loc}\\{self.process_handler.data_params.security}\\'
             f'{self.process_handler.data_params.time_frame_test}\\'
             f'{self.process_handler.data_params.time_frame_test}_test_{self.process_handler.data_params.time_len}\\'
             f'{self.process_handler.side}')

        self.data_folder = f'{self.param_folder}\\Data'
        self.model_folder = f'{self.param_folder}\\Models'
        self.main_train_path = f'{self.model_folder}\\{self.lstm_model.side}\\param_{self.lstm_model.param}_main_model'

        for folder in [self.param_folder, self.data_folder, self.model_folder]:
            os.makedirs(folder, exist_ok=True)

    def set_model_train_paths(self, test_date):
        self.check_create_model_folder()

        previous_test_date = pd.to_datetime(test_date) - timedelta(days=self.data_params.test_period_days)
        previous_test_date = previous_test_date.strftime(format='%Y-%m-%d')
        self.test_date = test_date

        self.model_save_path = \
            f'{self.model_folder}\\{self.lstm_model.side}_{self.lstm_model.param}\\param_{self.test_date}_model'
        self.previous_model_path = \
            f'{self.model_folder}\\{self.lstm_model.side}_{self.lstm_model.param}\\param_{previous_test_date}_model'

        # for folder in [self.model_save_path, self.previous_model_path]:
        #     os.makedirs(folder, exist_ok=True)

    def save_model(self, i):
        print(f'Model Saved: {self.model_save_path}')
        self.lstm_model.model.save(f'{self.model_save_path}\\model.keras')
        if i == 0:
            self.lstm_model.model.save(f'{self.main_train_path}\\model.keras')

    def load_prior_test_date_model(self):
        last_test_date = (pd.to_datetime(self.test_date, format='%Y-%m-%d') -
                          timedelta(days=self.data_params.test_period_days)).strftime(format='%Y-%m-%d')
        threshold = self.lstm_model.opt_threshold
        print(f'Loading Prior Week Model: {str(last_test_date)}')

        class_weights = self.lstm_model.get_class_weights()
        wce_accuracy = lf.weighted_categorical_crossentropy(class_weights)
        combined_wl_loss = lf.comb_focal_wce_f1(beta=2.0,
                                                opt_threshold=threshold,
                                                class_weights=class_weights)
        f1_score = lf.beta_f1()
        npv_fn = lf.negative_predictive_value(threshold)
        focal_loss_fn = lf.focal_loss()
        huber_loss = lf.weighted_huber_loss()
        auc = lf.weighted_auc(class_weights)

        self.lstm_model.model = (
            keras.models.load_model(f'{self.previous_model_path}\\model.keras',
                                    custom_objects={'focal_loss_fixed': focal_loss_fn,
                                                    'combined_wl_loss': combined_wl_loss,
                                                    'wce_loss': wce_accuracy,
                                                    'f1_score': f1_score,
                                                    'npv': npv_fn,
                                                    'bal_accuracy': lf.bal_accuracy,
                                                    'auc_loss': auc,
                                                    'recall': Recall(thresholds=threshold, name='recall'),
                                                    'huber_loss': huber_loss,
                                                    'TemperatureScalingLayer': TemperatureScalingLayer
                                                    }))

    def load_current_test_date_model(self):
        print(f'Loading Current Week Model: {self.test_date}')
        threshold = self.lstm_model.opt_threshold
        class_weights = self.lstm_model.get_class_weights()
        focal_loss_fn = lf.focal_loss()
        combined_wl_loss = lf.comb_focal_wce_f1(beta=2.0,
                                                opt_threshold=threshold,
                                                class_weights=class_weights)
        f1_score = lf.beta_f1()
        npv_fn = lf.negative_predictive_value(threshold)
        wce_accuracy = lf.weighted_categorical_crossentropy(class_weights)
        huber_loss = lf.weighted_huber_loss()
        auc = lf.weighted_auc(class_weights)

        self.lstm_model.model = (
            keras.models.load_model(f'{self.model_save_path}\\model.keras',
                                    custom_objects={'focal_loss_fixed': focal_loss_fn,
                                                    'combined_wl_loss': combined_wl_loss,
                                                    'wce_loss': wce_accuracy,
                                                    'f1_score': f1_score,
                                                    'npv': npv_fn,
                                                    'bal_accuracy': lf.bal_accuracy,
                                                    'auc_loss': auc,
                                                    'recall': Recall(thresholds=threshold, name='recall'),
                                                    'huber_loss': huber_loss,
                                                    'TemperatureScalingLayer': TemperatureScalingLayer
                                                    }))
        self.lstm_model.model.load_weights(f'{self.model_save_path}\\model.keras')

    def save_scalers(self):
        scalers = {
            'y_pnl_scaler': self.mkt_data.y_pnl_scaler,
            'intra_scaler': self.mkt_data.intra_scaler,
            'daily_scaler': self.mkt_data.daily_scaler
        }
        os.makedirs(os.path.dirname(f'{self.model_save_path}\\'), exist_ok=True)
        for key, val in scalers.items():
            with open(f'{self.model_save_path}\\{key}.pkl', 'wb') as f:
                pickle.dump(val, f)
            print(f'Saved {key} Scaler\n')

    def load_scalers(self, retrain=False):
        curr_scalers_exist = os.path.exists(f'{self.model_save_path}\\intra_scaler.pkl')
        prev_scalers_exist = os.path.exists(f'{self.previous_model_path}\\intra_scaler.pkl')

        if (not retrain or not curr_scalers_exist) and prev_scalers_exist:
            with open(f'{self.previous_model_path}\\y_pnl_scaler.pkl', 'rb') as f:
                self.mkt_data.y_pnl_scaler = pickle.load(f)
            with open(f'{self.previous_model_path}\\intra_scaler.pkl', 'rb') as f:
                self.mkt_data.intra_scaler = pickle.load(f)
            with open(f'{self.previous_model_path}\\daily_scaler.pkl', 'rb') as f:
                self.mkt_data.daily_scaler = pickle.load(f)
                print(f'...Loaded Previous Scalers: \n'
                      f'{self.previous_model_path}')

        elif curr_scalers_exist:
            with open(f'{self.model_save_path}\\y_pnl_scaler.pkl', 'rb') as f:
                self.mkt_data.y_pnl_scaler = pickle.load(f)
            with open(f'{self.model_save_path}\\intra_scaler.pkl', 'rb') as f:
                self.mkt_data.intra_scaler = pickle.load(f)
            with open(f'{self.model_save_path}\\daily_scaler.pkl', 'rb') as f:
                self.mkt_data.daily_scaler = pickle.load(f)
                print(f'Loaded Previous Scalers: \n'
                      f'...{self.model_save_path}')
        else:
            pass

    def save_all_prediction_data(self, side, param, test_date, model_dfs, trade_dfs):
        self.save_metrics(side, param, test_date, model_dfs, 'Model')
        self.save_metrics(side, param, test_date, trade_dfs[0], 'WL')
        self.save_metrics(side, param, test_date, trade_dfs[1], 'PnL')
        if self.process_handler.train_modeltf:
            self.save_plot_to_excel(side)

    def save_metrics(self, side, param, test_date, dfs, sheet_name, stack_row=False):
        os.makedirs(f'{self.data_folder}\\{side}_{param}', exist_ok=True)
        self.save_file = f'{self.data_folder}\\{side}_{param}\\predictions_{side}_{param}_{test_date}.xlsx'
        sheet_name = f'{side}_{sheet_name}'

        if os.path.exists(self.save_file):
            # Load the existing workbook
            book = openpyxl.load_workbook(self.save_file)
            if not book.sheetnames:
                book.create_sheet(sheet_name, -1)
                book.active.title = sheet_name

            with pd.ExcelWriter(self.save_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                start_positions = get_excel_sheet_df_positions(dfs, stack_row)
                write_metrics_to_excel(writer, dfs, sheet_name, start_positions)

        else:
            # Create a new Excel file
            create_new_excel_file(self.save_file, sheet_name)
            with pd.ExcelWriter(self.save_file, engine='openpyxl') as writer:
                start_positions = get_excel_sheet_df_positions(dfs, stack_row)
                write_metrics_to_excel(writer, dfs, sheet_name, start_positions)

    def save_plot_to_excel(self, side):
        file_exists = os.path.exists(self.save_file)
        if not file_exists:
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(self.save_file)

        # Select a sheet or create a new one
        sheet_name = f'{side}_Model'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        img_loc = f'{self.data_folder}\\temp_img.png'
        if self.lstm_model.model_plot:
            self.lstm_model.model_plot.fig.savefig(img_loc)
        img = Image(img_loc)

        plot_loc_excel = 'F2'
        if file_exists:
            plot_loc_excel = 'F12'
        ws.add_image(img, plot_loc_excel)

        wb.save(self.save_file)

        if os.path.exists(img_loc):
            os.remove(img_loc)


def write_metrics_to_excel(writer, dfs, sheet_name, start_positions):
    for df, (startrow, startcol) in zip(dfs, start_positions):
        df.to_excel(writer, sheet_name=sheet_name,
                    startrow=startrow, startcol=startcol)


def get_excel_sheet_df_positions(dfs, stack_row):
    start_positions = [(0, 0)]

    if len(dfs) > 1:
        if stack_row:
            start_row = len(dfs[0])
            for df in dfs[1:]:
                start_row += 2
                start_positions.append((start_row, 0))
                start_row += len(df)
        else:
            start_row = 0
            start_col = len(dfs[0].columns) + 2
            for df in dfs[1:]:
                start_positions.append((start_row, start_col))
                start_row += len(df) + 2

    return start_positions


def create_new_excel_file(file_path, sheet_name):
    if not os.path.exists(file_path):
        df = pd.DataFrame()
        df.to_excel(file_path, sheet_name=sheet_name)

