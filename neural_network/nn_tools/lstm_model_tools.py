import random
import time
import numpy as np
import pandas as pd
import os
import io
import tensorflow as tf
from tensorflow.keras.models import load_model, Model
from tensorflow.keras.layers import LSTM, Dense, Dropout, Input, Concatenate, BatchNormalization
from tensorflow.keras.optimizers import Adam
from keras.regularizers import l2, l1
from tensorflow.keras.initializers import GlorotUniform
from tensorflow.keras.callbacks import ReduceLROnPlateau, Callback, EarlyStopping
from sklearn.utils.class_weight import compute_class_weight
import matplotlib.pyplot as plt
import neural_network.nn_tools.math_tools as mt
from openpyxl import load_workbook
import neural_network.nn_tools.lstm_data_tools as ldt
from openpyxl.drawing.image import Image
import openpyxl


class LstmOptModel:
    def __init__(self, lstm_dict):
        self.epochs = lstm_dict['epochs']
        self.batch_s = lstm_dict['batch_size']
        self.test_size = lstm_dict['test_size']
        self.max_acc = lstm_dict['max_accuracy']
        self.input_layer_daily = None
        self.input_layer_intraday = None
        self.win_loss_output = None
        self.float_output = None

        self.daily_len = 23
        self.intra_len = 24

        self.model = None
        self.scheduler = None
        self.optimizer = None
        self.input_shapes = None

        self.model_summary = None
        self.model_plot = None
        self.model_save_path = None

    def check_for_previous_model(self, lstm_data, side, param):
        self.model_save_path = f'{lstm_data.trade_data.data_loc}\\Results\\{side}_{param}_model.h5'
        previous_train = os.path.exists(self.model_save_path)

        return previous_train

    def set_scheduler(self):
        self.scheduler = CustLRSchedule(initial_learning_rate=.03, decay_rate=0.00001)

    def set_intra_len(self, intra_len):
        self.intra_len = intra_len

    def build_compile_model(self, lstm_data):
        print('\nBuilding New Model')
        self.set_scheduler()
        self.build_lstm_model(lstm_data.dailydata, lstm_data.intradata)
        self.compile_model(lstm_data)

    def compile_model(self, lstm_data):
        self.optimizer = Adam(.0005)

        class_weights = compute_class_weight('balanced',
                                             classes=np.unique(lstm_data.y_train_wl_df['Win']),
                                             y=lstm_data.y_train_wl_df['Win'])
        class_weight_dict_wl = {i: weight for i, weight in enumerate(class_weights)}
        print(f'Class Weights: {class_weight_dict_wl}\n')

        loss_fn_wl = weighted_categorical_crossentropy(class_weights)

        self.model = Model(inputs=[self.input_layer_daily, self.input_layer_intraday],
                           outputs=[self.win_loss_output, self.float_output])
        self.model.compile(optimizer=self.optimizer,
                           loss={'wl_classification': loss_fn_wl,
                                 'pnl_output': 'mse'},
                           metrics={'wl_classification': 'accuracy',
                                    'pnl_output': 'mse'})
        print('New Model Created')
        self.get_model_summary_df()

    def get_input_shapes(self, daily_data, intraday_data):
        """Plus 12 for number of months in year"""
        daily_shape = (self.daily_len, len(daily_data.columns) - 1)
        intraday_shape = (self.intra_len, len(intraday_data.columns) - 1)

        self.input_shapes = [daily_shape, intraday_shape]

    def build_lstm_model(self, daily_data, intraday_data):
        self.get_input_shapes(daily_data, intraday_data)

        # Daily LSTM branch
        self.input_layer_daily = Input(self.input_shapes[0])
        lstm_d1 = LSTM(units=48,
                       activation='tanh',
                       recurrent_activation='sigmoid',
                       return_sequences=True,  # Keep this to pass the sequence
                       kernel_initializer=GlorotUniform(),
                       kernel_regularizer=l2(0.005),
                       name='lstm_d1')(self.input_layer_daily)

        batch_n1 = BatchNormalization(name='batch_n1')(lstm_d1)

        drop_d1 = Dropout(0.05, name='drop_d1')(batch_n1)

        # Final LSTM layer for daily data - set return_sequences=False
        lstm_d2 = LSTM(units=32,
                       activation='tanh',
                       recurrent_activation='sigmoid',
                       return_sequences=False,
                       kernel_initializer=GlorotUniform(),
                       kernel_regularizer=l2(0.005),
                       name='lstm_d2')(drop_d1)

        dense_d1 = Dense(units=32,
                         activation='tanh',
                         kernel_initializer=GlorotUniform(),
                         kernel_regularizer=l2(0.005),
                         name='dense_d1')(lstm_d2)

        # Intraday LSTM branch
        self.input_layer_intraday = Input(self.input_shapes[1])
        lstm_i1 = LSTM(units=192,
                       activation='tanh',
                       recurrent_activation='sigmoid',
                       return_sequences=True,
                       kernel_initializer=GlorotUniform(),
                       kernel_regularizer=l2(0.005),
                       name='lstm_i1')(self.input_layer_intraday)

        batch_n2 = BatchNormalization(name='batch_n2')(lstm_i1)

        drop_i1 = Dropout(0.05, name='drop_i1')(batch_n2)

        # Final LSTM layer for intraday data - set return_sequences=False
        lstm_i2 = LSTM(units=96,
                       activation='tanh',
                       recurrent_activation='sigmoid',
                       return_sequences=False,
                       kernel_initializer=GlorotUniform(),
                       kernel_regularizer=l2(0.005),
                       name='lstm_i2')(drop_i1)

        dense_i1 = Dense(units=64,
                         activation='tanh',
                         kernel_initializer=GlorotUniform(),
                         kernel_regularizer=l2(0.005),
                         name='dense_i1')(lstm_i2)

        # Combine the branches
        merged_lstm = Concatenate()([dense_d1, dense_i1])

        drop_m1 = Dropout(0.05, name='drop_m1')(merged_lstm)

        dense_m1 = Dense(units=32,
                         activation='tanh',
                         kernel_initializer=GlorotUniform(),
                         kernel_regularizer=l2(0.005),
                         name='dense_m1')(drop_m1)

        # Output layers
        self.win_loss_output = Dense(2,
                                     activation='softmax',
                                     name='wl_classification')(dense_m1)

        self.float_output = Dense(units=1,
                                  activation='tanh',
                                  name='pnl_output')(dense_m1)

    def train_model(self, lstm_data):
        lr_scheduler = ReduceLROnPlateau(monitor='loss', factor=0.75, patience=2, min_lr=.00001, verbose=1)
        self.model_plot = LivePlotLosses()
        data_gen = CustomDataGenerator(lstm_data, self, self.batch_s)
        stop_at_accuracy = StopAtAccuracy(accuracy_threshold=self.max_acc)
        self.model.fit(data_gen,
                       epochs=self.epochs,
                       verbose=1,
                       callbacks=[lr_scheduler, self.model_plot, stop_at_accuracy])

    def evaluate_model(self, lstm_data):
        test_generator = CustomDataGenerator(lstm_data, self, self.batch_s, train=False)

        # Evaluate the model on the test data
        test_loss, test_wl_loss, test_pnl_loss, wl_accuracy, pnl_output_mse = self.model.evaluate(test_generator)

        print(f'Test Loss: {test_loss:.4f}')
        print(f'Win/Loss Classification Loss: {test_wl_loss:.4f}')
        print(f'PNL Output Loss: {test_pnl_loss:.4f}')
        print(f'WL Classification Acc: {wl_accuracy:.4f}')
        print(f'PnL Output MSE: {pnl_output_mse:.4f}')

        return test_loss, test_wl_loss, test_pnl_loss, wl_accuracy, pnl_output_mse

    def predict_data_evaluate(self, lstm_data, param, side):
        save_handler = ldt.SaveHandler(lstm_data)
        model_metrics = self.evaluate_model(lstm_data)
        model_metrics = pd.DataFrame(model_metrics).T
        model_metrics.columns = ['test_loss', 'test_wl_loss', 'test_pnl_loss', 'wl_accuracy', 'pnl_output_mse']

        test_generator = CustomDataGenerator(lstm_data, self, self.batch_s, train=False)

        wl_predictions, pnl_predictions = self.model.predict(test_generator, verbose=1)
        lstm_data.prep_predicted_data(wl_predictions, pnl_predictions)

        wl_con_mat = lstm_data.get_confusion_matrix_metrics()
        save_handler.get_trade_metrics()
        wl_dfs = [lstm_data.y_test_wl_df, save_handler.trade_metrics] + [pd.DataFrame(l) for l in wl_con_mat]
        save_handler.save_metrics(side, param, wl_dfs, 'WL')

        pnl_con_mat = lstm_data.get_confusion_matrix_metrics(wl=False)
        save_handler.get_trade_metrics(wl=False)
        pnl_dfs = [lstm_data.y_test_pnl_df, save_handler.trade_metrics] + [pd.DataFrame(l) for l in pnl_con_mat]
        save_handler.save_metrics(side, param, pnl_dfs, 'PnL')

        model_dfs = [self.model_summary, model_metrics]
        save_handler.save_metrics(side, param, model_dfs, 'Model')

        self.save_plot_to_excel(save_handler, side)

        print(f'Data saved to: {save_handler.save_file}')

        model_save_loc = f'{save_handler.save_loc}\\{side}_{param}_model.h5'
        self.model.save(model_save_loc)
        print(f'Saved model to: {model_save_loc}')
        # breakpoint()

    def save_plot_to_excel(self, save_handler, side):
        file_exists = os.path.exists(save_handler.save_file)
        if not file_exists:
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(save_handler.save_file)

        # Select a sheet or create a new one
        sheet_name = f'{side}_Model'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        img_loc = f'{save_handler.save_loc}\\temp_img.png'
        self.model_plot.fig.savefig(img_loc)
        img = Image(img_loc)

        plot_loc_excel = 'F2'
        if file_exists:
            plot_loc_excel = 'F12'
        ws.add_image(img, plot_loc_excel)

        wb.save(save_handler.save_file)

        if os.path.exists(img_loc):
            os.remove(img_loc)

    def get_model_summary_df(self):
        self.model.summary()

        summary_buf = io.StringIO()
        self.model.summary(print_fn=lambda x: summary_buf.write(x + "\n"))

        summary_string = summary_buf.getvalue()
        summary_lines = summary_string.split("\n")

        summary_data = []
        for line in summary_lines:
            split_line = list(filter(None, line.split(" ")))
            if len(split_line) > 1:
                summary_data.append(split_line)

        # df_summary = pd.DataFrame(summary_data, columns=['Layer', 'Output Shape', 'Param #', 'Connected To'])
        df_summary = pd.DataFrame(summary_data)
        df_cols = df_summary.iloc[1]
        df_summary = df_summary.iloc[2:].reset_index(drop=True)
        df_summary.columns = df_cols

        self.model_summary = df_summary


"""--------------------------------------------Custom Callbacks Work-------------------------------------------------"""


class CustLRSchedule(tf.keras.optimizers.schedules.LearningRateSchedule):
    def __init__(self, initial_learning_rate, decay_rate):
        self.initial_learning_rate = initial_learning_rate
        self.decay_rate = decay_rate

    def __call__(self, step):
        # Define the custom learning rate schedule here.
        # For example, a simple exponential decay:
        return self.initial_learning_rate / (1 + self.decay_rate * step)

    def get_config(self):
        # Required if you need to save the schedule.
        return {
            'initial_learning_rate': self.initial_learning_rate,
            'decay_rate': self.decay_rate
        }


"""--------------------------------------------Custom Callbacks Work-------------------------------------------------"""


class CustomDataGenerator(tf.keras.utils.Sequence):
    def __init__(self, lstm_data, lstm_model, batch_size, train=True):
        self.lstm_data = lstm_data
        self.lstm_model = lstm_model
        self.train_tf = train
        self.sample_ind_list = []
        self.n_samples = 0
        self.batch_size = batch_size
        self.set_attributes()

    def __len__(self):
        # Number of batches per epoch
        return int(np.ceil(self.n_samples / self.batch_size))

    def __getitem__(self, index):
        if index == 0:
            np.random.shuffle(self.sample_ind_list)

        start_ind = index * self.batch_size
        end_ind = (index + 1) * self.batch_size
        train_inds = self.sample_ind_list[start_ind:end_ind]
        batch_gen = self.lstm_data.create_batch_input(train_inds,
                                                      self.lstm_model.daily_len,
                                                      self.lstm_model.intra_len,
                                                      self.train_tf)

        x_day_arr, x_intra_arr, y_pnl_arr, y_wl_arr = next(batch_gen)

        return [x_day_arr, x_intra_arr], {'wl_classification': y_wl_arr, 'pnl_output': y_pnl_arr}

    def set_attributes(self):
        if self.train_tf:
            self.sample_ind_list = list(self.lstm_data.y_train_pnl_df.index)
            self.n_samples = len(self.lstm_data.trade_data.y_train_df)

        else:
            self.sample_ind_list = list(self.lstm_data.y_test_pnl_df.index)
            self.n_samples = len(self.lstm_data.trade_data.y_test_df)

    def on_epoch_end(self):
        if self.train_tf:
            np.random.shuffle(self.sample_ind_list)


def weighted_categorical_crossentropy(class_weights):
    def loss(y_true, y_pred):
        y_true = tf.cast(y_true, tf.float32)
        weights = tf.reduce_sum(class_weights * y_true, axis=-1)
        return tf.reduce_mean(weights * tf.keras.losses.categorical_crossentropy(y_true, y_pred))

    return loss


class LivePlotLosses(Callback):
    def __init__(self):
        super(LivePlotLosses, self).__init__()
        self.epochs = []
        self.losses = []
        self.wl_classification_losses = []
        self.pnl_output_losses = []
        self.wl_classification_accuracies = []
        self.pnl_output_mses = []

        plt.ion()  # Interactive mode on for live plotting
        self.fig, self.axs = plt.subplots(2, 2, figsize=(10, 6))  # Create a 2x2 grid of subplots
        self.fig.tight_layout()

    def on_epoch_end(self, epoch, logs=None):
        logs = logs or {}
        self.epochs.append(epoch + 1)  # Accumulate epochs
        self.losses.append(logs.get('loss'))
        self.wl_classification_losses.append(logs.get('wl_classification_loss'))
        self.pnl_output_losses.append(logs.get('pnl_output_loss'))
        self.wl_classification_accuracies.append(logs.get('wl_classification_accuracy'))
        self.pnl_output_mses.append(logs.get('pnl_output_mse'))

        # Clear previous plots
        for ax in self.axs.flat:
            ax.clear()

        # Plot accumulated data for all completed epochs
        self.axs[0, 0].plot(self.epochs, self.losses,
                            label="Total Loss", marker='o')
        self.axs[0, 0].set_title("Total Loss")
        self.axs[0, 0].legend()

        self.axs[0, 1].plot(self.epochs, self.wl_classification_losses,
                            label="WL Classification Loss", marker='o')
        self.axs[0, 1].set_title("WL Classification Loss")
        self.axs[0, 1].legend()

        self.axs[1, 0].plot(self.epochs, self.wl_classification_accuracies,
                            label="WL Classification Accuracy", marker='o')
        self.axs[1, 0].set_title("WL Classification Accuracy")
        self.axs[1, 0].legend()

        self.axs[1, 1].plot(self.epochs, self.pnl_output_mses,
                            label="PnL Output MSE", marker='o')
        self.axs[1, 1].set_title("PnL Output MSE")
        self.axs[1, 1].legend()

        # Draw the updated plots and pause for a short moment to update the plot
        self.fig.canvas.draw()
        plt.pause(0.2)  # Pause briefly to ensure the plot refreshes

    def on_train_end(self, logs=None):
        plt.ioff()  # Turn off interactive mode at the end
        plt.close()


class StopAtAccuracy(Callback):
    def __init__(self, accuracy_threshold=0.95):
        super(StopAtAccuracy, self).__init__()
        self.accuracy_threshold = accuracy_threshold

    def on_epoch_end(self, epoch, logs=None):
        wl_accuracy = logs.get('wl_classification_accuracy')
        if wl_accuracy is not None and wl_accuracy >= self.accuracy_threshold:
            print(f"\nReached {self.accuracy_threshold*100}% accuracy, stopping training!")
            self.model.stop_training = True
