import numpy as np
import os
import pandas as pd
import old_files.math_tools as mt
import data_tools.general_tools as gt
import data_tools.data_trade_tools as tdt
from sklearn.preprocessing import StandardScaler, OneHotEncoder
from sklearn.metrics import confusion_matrix, classification_report
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook


class LstmDataHandler:
    def __init__(self, data_param_dict):
        self.data_params = data_param_dict
        self.all_secs = [self.data_params['security']] + self.data_params['other_securities']
        self.trade_data = tdt.TradeData

        self.dailydata = None
        self.intradata = None
        self.security_df = None

        self.intra_scaler = None
        self.daily_scaler = None
        self.month_onehot_scaler = None
        self.day_onehot_scaler = None

        self.x_train_intra = None
        self.x_test_intra = None

        self.y_train_pnl_df = None
        self.y_train_wl_df = None
        self.y_test_pnl_df = None
        self.y_test_wl_df = None

        self.y_pnl_scaler = None
        self.y_wl_onehot_scaler = None

        self.intra_len = 0
        self.wl_ratio = 0

    def set_daily_time_len(self):
        open_time = datetime.strptime(f'{self.data_params["start_hour"]}:00', '%H:%M')
        close_time = datetime.strptime('15:00', '%H:%M')
        time_diff = close_time - open_time
        time_interval = int(''.join(filter(str.isdigit, self.data_params['time_frame'])))
        self.intra_len = int(time_diff/timedelta(minutes=time_interval))

    def load_prep_daily_data(self):
        print('\nLoading Daily Data')
        daily_dfs = []
        for sec in [self.data_params['security']] + self.data_params['other_securities']:
            print(f'...{sec}')
            temp_df = pd.read_csv(f'{self.data_params["data_loc"]}\\{sec}_daily_20240505_20040401.txt')
            temp_df['ATR'] = mt.create_atr(temp_df, 8)
            temp_df = gt.convert_date_time(temp_df)
            temp_df = temp_df.sort_values(by='DateTime')
            for col in temp_df.columns:
                if col in ['Vol.1', 'OI', 'Time', 'AvgExp12', 'AvgExp24', 'Bearish_Double_Candle',
                           'Bullish_Double_Candle', 'DateTime']:
                    temp_df.drop(columns=[col], inplace=True)

            temp_df = mt.set_various_data(temp_df)

            for col in temp_df.columns[1:]:
                temp_df[col] = temp_df[col].astype(np.float32)
                temp_df.rename(columns={col: f'{sec}_{col}'}, inplace=True)

            daily_dfs.append(temp_df)

        self.dailydata = daily_dfs[0]
        for df in daily_dfs[1:]:
            self.dailydata = pd.merge(self.dailydata, df, on='Date')

        self.dailydata = (
            self.dailydata)[self.dailydata['Date'] >= pd.to_datetime(self.data_params['start_train_date']).date()]
        self.dailydata['Date'] = pd.to_datetime(self.dailydata['Date'])
        self.dailydata['Month'] = self.dailydata['Date'].dt.month
        self.dailydata['Day'] = self.dailydata['Date'].dt.dayofweek
        self.dailydata['Date'] = self.dailydata['Date'].dt.date

        self.finish_data_prep()

    def load_prep_intra_data(self):
        print('\nLoading Intraday Data')
        intra_dfs = []
        for sec in [self.data_params['security']] + self.data_params['other_securities']:
            print(f'...{sec}')
            temp_df = pd.read_csv(f'{self.data_params["data_loc"]}\\'
                                  f'{sec}_{self.data_params["time_frame"]}_20240505_20040401.txt')
            temp_df['ATR'] = mt.create_atr(temp_df, 8)
            temp_df = gt.convert_date_time(temp_df)
            temp_df = temp_df.sort_values(by='DateTime')
            if sec == self.data_params['security']:
                self.security_df = temp_df

            for col in temp_df.columns:
                if col in ['Vol.1', 'OI', 'Time', 'AvgExp12', 'AvgExp24', 'Bearish_Double_Candle',
                           'Bullish_Double_Candle', 'Date', 'Time', 'Up', 'Down']:
                    temp_df.drop(columns=[col], inplace=True)

            temp_df = temp_df[['DateTime'] + temp_df.columns[:-1].to_list()]
            temp_df = mt.set_various_data(temp_df)

            for col in temp_df.columns[1:]:
                temp_df[col] = temp_df[col].astype(np.float32)
                temp_df.rename(columns={col: f'{sec}_{col}'}, inplace=True)

            intra_dfs.append(temp_df)

        self.intradata = intra_dfs[0]
        for df in intra_dfs[1:]:
            self.intradata = pd.merge(self.intradata, df, on='DateTime')

        self.intradata = (
            self.intradata)[self.intradata['DateTime'].dt.date >=
                            pd.to_datetime(self.data_params['start_train_date']).date()]
        self.intradata['DateTime'] = pd.to_datetime(self.intradata['DateTime'])
        self.intradata['Month'] = self.intradata['DateTime'].dt.month
        self.intradata['Day'] = self.intradata['DateTime'].dt.dayofweek

        self.finish_data_prep(daily=False)

    def finish_data_prep(self, daily=True):
        print('\nFinishing Intraday Data Prep')
        if daily:
            df = self.dailydata
        else:
            df = self.intradata

        df = mt.create_rsi(df, self.all_secs)
        df = mt.add_high_low_diff(df,
                                  self.data_params['other_securities'],
                                  self.data_params['security'])

        df = mt.smooth_vol_oi(df, self.all_secs)
        df = mt.scale_open_close(df)
        df = gt.sort_data_cols(df)
        df = gt.fill_na_inf(df)

        if daily:
            self.dailydata = df
        else:
            self.intradata = df

    def set_x_train_test_datasets(self):
        print('\nBuilding X-Train and Test Datasets')
        self.x_train_intra = self.intradata[self.intradata['DateTime'].dt.date.isin(self.trade_data.train_dates)]
        self.x_train_intra = gt.fill_na_inf(self.x_train_intra)

        self.x_test_intra = self.intradata[self.intradata['DateTime'].dt.date.isin(self.trade_data.test_dates)]
        self.x_test_intra = gt.fill_na_inf(self.x_test_intra)

    def scale_x_data(self):
        print('\nScaling X Data')
        # self.intra_scaler = RobustScaler(quantile_range=(3, 97))
        self.x_train_intra = gt.arrange_xcols_for_scaling(self.x_train_intra)
        self.x_test_intra = gt.arrange_xcols_for_scaling(self.x_test_intra)
        self.intra_scaler = StandardScaler()
        self.x_train_intra.iloc[:, 1:] = (
            self.intra_scaler.fit_transform(self.x_train_intra.iloc[:, 1:].values.astype('float32')))
        self.x_test_intra.iloc[:, 1:] = (
            self.intra_scaler.transform(self.x_test_intra.iloc[:, 1:].values.astype('float32')))

        # self.daily_scaler = RobustScaler()
        self.dailydata = gt.arrange_xcols_for_scaling(self.dailydata)
        self.daily_scaler = StandardScaler()
        self.dailydata.iloc[:, 1:] = (
            self.intra_scaler.fit_transform(self.dailydata.iloc[:, 1:].values.astype('float32')))

        # self.onehot_month_day()

    def onehot_month_day(self):
        self.day_onehot_scaler = OneHotEncoder(sparse_output=False)
        oh_day_train = self.day_onehot_scaler.fit_transform(self.x_train_intra.iloc[:, -2].values.reshape(-1, 1))
        oh_day_test = self.day_onehot_scaler.transform(self.x_test_intra.iloc[:, -2].values.reshape(-1, 1))
        oh_day_daily = self.day_onehot_scaler.transform(self.dailydata.iloc[:, -2].values.reshape(-1, 1))

        self.month_onehot_scaler = OneHotEncoder(sparse_output=False)
        oh_month_train = self.day_onehot_scaler.fit_transform(self.x_train_intra.iloc[:, -1].values.reshape(-1, 1))
        oh_month_test = self.day_onehot_scaler.transform(self.x_test_intra.iloc[:, -1].values.reshape(-1, 1))
        oh_month_daily = self.day_onehot_scaler.transform(self.dailydata.iloc[:, -1].values.reshape(-1, 1))

        self.x_train_intra = self.x_train_intra.iloc[:, :-2]
        self.x_test_intra = self.x_test_intra.iloc[:, :-2]
        self.dailydata = self.dailydata.iloc[:, :-2]

        self.x_train_intra = gt.add_arr_to_df(self.x_train_intra, oh_day_train)
        self.x_test_intra = gt.add_arr_to_df(self.x_test_intra, oh_day_test)
        self.dailydata = gt.add_arr_to_df(self.dailydata, oh_day_daily)

        self.x_train_intra = gt.add_arr_to_df(self.x_train_intra, oh_month_train)
        self.x_test_intra = gt.add_arr_to_df(self.x_test_intra, oh_month_test)
        self.dailydata = gt.add_arr_to_df(self.dailydata, oh_month_daily)

    def scale_y_pnl_data(self):
        print('\nScaling Y-pnl Data')
        # self.y_pnl_scaler = RobustScaler(quantile_range=(5, 95))
        self.y_pnl_scaler = StandardScaler()
        self.y_train_pnl_df = self.trade_data.y_train_df.iloc[:, :2]
        pnl_scaled = (
            self.y_pnl_scaler.fit_transform(self.y_train_pnl_df.iloc[:, 1].values.astype('float32').reshape(-1, 1)))
        self.y_train_pnl_df.iloc[:, 1] = pnl_scaled.reshape(-1, 1)

        self.y_test_pnl_df = self.trade_data.y_test_df.iloc[:, :2]
        pnl_scaled = (
            self.y_pnl_scaler.transform(self.y_test_pnl_df.iloc[:, 1].values.astype('float32').reshape(-1, 1)))
        self.y_test_pnl_df.iloc[:, 1] = pnl_scaled.reshape(-1, 1)

    def onehot_y_wl_data(self):
        print('\nOnehotting WL Data')
        self.y_wl_onehot_scaler = OneHotEncoder(sparse_output=False)

        self.y_train_wl_df = self.trade_data.y_train_df.iloc[:, [0, 2]]
        wl_dat = self.y_train_wl_df.iloc[:, 1].values
        wl_dat = self.y_wl_onehot_scaler.fit_transform(wl_dat.reshape(-1, 1))

        self.y_train_wl_df['Loss'] = wl_dat[:, 0]
        self.y_train_wl_df['Win'] = wl_dat[:, 1]
        self.y_train_wl_df.drop('Win_Loss', inplace=True, axis=1)

        self.y_test_wl_df = self.trade_data.y_test_df.iloc[:, [0, 2]]
        wl_dat = self.y_test_wl_df.iloc[:, 1].values
        wl_dat = self.y_wl_onehot_scaler.transform(wl_dat.reshape(-1, 1))

        self.y_test_wl_df['Loss'] = wl_dat[:, 0]
        self.y_test_wl_df['Win'] = wl_dat[:, 1]
        self.y_test_wl_df.drop('Win_Loss', inplace=True, axis=1)

    def grab_prep_trade(self, y_pnl_df, y_wl_df, x_intraday, train_ind, daily_len, intra_len):
        while True:
            try:
                trade_dt = y_pnl_df.loc[train_ind, 'DateTime']

                x_daily_input = self.dailydata[self.dailydata['Date'] < trade_dt.date()].tail(daily_len).values[:, 1:]
                x_daily_input = gt.pad_to_length(x_daily_input, daily_len)

                x_intra_input = x_intraday[(x_intraday['DateTime'] <= trade_dt) &
                                           (x_intraday['DateTime'] >=
                                            trade_dt.replace(hour=self.data_params['start_hour'],
                                                             minute=self.data_params['start_minute']))]

                x_intra_input = x_intra_input.tail(intra_len).values[:, 1:]
                x_intra_input = gt.pad_to_length(x_intra_input, intra_len)

                y_pnl_input = np.array([y_pnl_df[y_pnl_df['DateTime'] == trade_dt].values[0, 1]])

                y_wl_input = y_wl_df[y_wl_df['DateTime'] == trade_dt].values[0, 1:]

                yield x_daily_input, x_intra_input, y_pnl_input, y_wl_input

            except StopIteration:
                break

    def create_batch_input(self, train_inds, daily_len, intra_len, train=True):
        while True:
            if train:
                y_pnl_df = self.y_train_pnl_df
                y_wl_df = self.y_train_wl_df
                x_intraday = self.x_train_intra
            else:
                y_pnl_df = self.y_test_pnl_df
                y_wl_df = self.y_test_wl_df
                x_intraday = self.x_test_intra

            x_day_arr, x_intra_arr, y_pnl_arr, y_wl_arr = [], [], [], []

            try:
                for train_ind in train_inds:
                    x_day, x_intra, y_pnl, y_wl = next(self.grab_prep_trade(y_pnl_df, y_wl_df, x_intraday,
                                                                            train_ind, daily_len, intra_len))
                    x_day_arr.append(x_day)
                    x_intra_arr.append(x_intra)
                    y_pnl_arr.append(y_pnl)
                    y_wl_arr.append(y_wl)

                x_day_arr = np.array(x_day_arr).astype(np.float32)
                x_intra_arr = np.array(x_intra_arr).astype(np.float32)
                y_pnl_arr = np.array(y_pnl_arr).astype(np.float32)
                y_wl_arr = np.array(y_wl_arr).astype(np.float32)

                yield x_day_arr, x_intra_arr, y_pnl_arr, y_wl_arr

                x_day_arr, x_intra_arr, y_pnl_arr, y_wl_arr = [], [], [], []

            except StopIteration:
                if x_day_arr:
                    x_day_arr = np.array(x_day_arr).astype(np.float32)
                    x_intra_arr = np.array(x_intra_arr).astype(np.float32)
                    y_pnl_arr = np.array(y_pnl_arr).astype(np.float32)
                    y_wl_arr = np.array(y_wl_arr).astype(np.float32)

                yield x_day_arr, x_intra_arr, y_pnl_arr, y_wl_arr

            break

    def win_loss_information(self):
        wins = sum(self.y_train_wl_df['Win'])
        losses = sum(self.y_train_wl_df['Loss'])
        self.wl_ratio = 1 - wins/(wins + losses)
        print(f'\nWin-Loss Ratio to Beat: {self.wl_ratio:.4f}')

    def add_close_to_test_dfs(self):
        self.y_test_pnl_df = self.y_test_pnl_df.merge(self.security_df[['DateTime', 'Close']], on=['DateTime'])
        self.y_test_wl_df = self.y_test_wl_df.merge(self.security_df[['DateTime', 'Close']], on=['DateTime'])

    def prep_predicted_data(self, wl_predictions, pnl_predictions):
        wl_predictions = self.y_wl_onehot_scaler.inverse_transform(wl_predictions)
        pnl_predictions = self.y_pnl_scaler.inverse_transform(pnl_predictions)
        self.y_test_pnl_df['PnL'] = (
            self.y_pnl_scaler.inverse_transform(self.y_test_pnl_df['PnL'].values.reshape(-1, 1)))

        self.y_test_wl_df['Pred'] = wl_predictions[:, 0]

        self.y_test_wl_df = (
            self.y_test_pnl_df[['DateTime', 'PnL']].merge(self.y_test_wl_df, on='DateTime'))
        self.y_test_pnl_df['Pred'] = pnl_predictions[:, 0]

        self.add_close_to_test_dfs()

        self.y_test_pnl_df = mt.summary_predicted(self.y_test_pnl_df)
        self.y_test_wl_df = mt.summary_predicted(self.y_test_wl_df, wl=True)

        self.prep_actual_wl_cols()
        self.prep_actual_wl_cols(wl=False)

    def prep_actual_wl_cols(self, wl=True):
        if wl:
            actual_wl = self.y_test_wl_df.apply(lambda row: 'Win' if row['PnL'] > 0 else 'Loss', axis=1)
            self.y_test_wl_df['Loss'] = actual_wl
            self.y_test_wl_df.drop(columns='Win', inplace=True)
            self.y_test_wl_df.rename(columns={'Loss': 'Actual_wl', 'Pred': 'Pred_wl'}, inplace=True)

        else:
            actual_wl = (
                self.y_test_pnl_df.apply(lambda row: 'Win' if row['PnL'] > 0 else 'Loss', axis=1))
            self.y_test_pnl_df['Actual_wl'] = actual_wl

            cols = list(self.y_test_pnl_df.columns)
            cols.insert(2, cols.pop(cols.index('Actual_wl')))
            self.y_test_pnl_df = self.y_test_pnl_df[cols]

            pred_wl = self.y_test_pnl_df.apply(lambda row: 'Win' if row['Pred'] > 0 else 'Loss', axis=1)
            self.y_test_pnl_df['Pred_wl'] = pred_wl

            cols = list(self.y_test_pnl_df.columns)
            cols.insert(4, cols.pop(cols.index('Pred_wl')))
            self.y_test_pnl_df = self.y_test_pnl_df[cols]

    def get_confusion_matrix_metrics(self, wl=True):
        self.win_loss_information()
        if wl:
            df = self.y_test_wl_df
            print('WL Dataset')
        else:
            df = self.y_test_pnl_df
            print('PnL Dataset')

        conf_matrix = confusion_matrix(df['Actual_wl'], df['Pred_wl'], labels=['Win', 'Loss'])
        conf_matrix = pd.DataFrame(conf_matrix, columns=['Pred_Pos', 'Pred_Neg'], index=['Actual_Pos', 'Actual_Neg'])
        conf_matrix = pd.DataFrame(conf_matrix).transpose()
        print('\nConfusion Matrix')
        print(conf_matrix)

        class_report = classification_report(df['Actual_wl'], df['Pred_wl'], target_names=['Win', 'Loss'],
                                             output_dict=True)
        print("\nClassification Report")
        print(class_report)

        return conf_matrix, class_report


class SaveHandler:
    def __init__(self, lstm_data):
        self.lstm_data = lstm_data
        self.save_loc = self.get_save_loc()
        self.save_file = None
        self.trade_metrics = None

        self.start_date = None
        self.end_date = None

    def get_save_loc(self):
        sec = self.lstm_data.data_params['security']
        timeframe = self.lstm_data.data_params['time_frame']
        self.start_date = self.lstm_data.data_params['start_train_date']
        self.end_date = self.lstm_data.data_params['final_train_date']

        save_loc = self.lstm_data.data_params['trade_dat_loc']

        save_loc = f'{save_loc}\\{sec}\\{timeframe}\\{timeframe}_test_20years\\Results'
        os.makedirs(save_loc, exist_ok=True)

        return save_loc

    def get_trade_metrics(self, wl=True):
        self.lstm_data.win_loss_information()
        if wl:
            df = self.lstm_data.y_test_wl_df
        else:
            df = self.lstm_data.y_test_pnl_df

        od_correct = (df['Pred_PnL'] > 0).sum()
        od_tot = (df['Pred_PnL'] != 0).sum()
        od_percent = od_correct/od_tot
        od_rolling = df['Pred_PnL_Total'].values
        od_max = np.max(od_rolling)
        od_avg = np.mean(df['Pred_PnL'].values)
        od_avg_win = np.mean(df.loc[df['Pred_PnL'] > 0, 'Pred_PnL'])
        od_avg_loss = np.mean(df.loc[df['Pred_PnL'] < 0, 'Pred_PnL'])
        od_max_draw = np.min(df['Pred_MaxDraw'].values)

        td_correct = (len(df[(df['Pred_wl'] == "Win") & (df['PnL'] > 0)]) +
                      len(df[(df['Pred_wl'] == "Loss") & (df['PnL'] < 0)]))
        td_tot = len(df.index)
        td_perecent = td_correct/td_tot
        df['Two_Dir_Pred_PnL'] = df.apply(
            lambda row: abs(row['PnL']) if
            (row['Pred_wl'] == 'Win' and row['PnL'] > 0) or
            (row['Pred_wl'] == 'Loss' and row['PnL'] < 0) else -abs(row['PnL']), axis=1)

        td_rolling = pd.Series(df['Two_Dir_Pred_PnL']).cumsum().values
        df['Two_Dir_Pred_Pnl_Total'] = td_rolling
        td_max = np.max(td_rolling)
        td_avg = np.mean(df['Two_Dir_Pred_PnL'].values)
        td_avg_win = np.mean(df.loc[df['Two_Dir_Pred_PnL'] > 0, 'Two_Dir_Pred_PnL'])
        td_avg_loss = np.mean(df.loc[df['Two_Dir_Pred_PnL'] < 0, 'Two_Dir_Pred_PnL'])
        df['Two_Dir_Pred_MaxDraw'] = mt.calculate_max_drawdown(df['Two_Dir_Pred_PnL'])
        td_max_draw = np.min(df['Two_Dir_Pred_MaxDraw'].values)

        od_dat = [od_correct, od_tot, od_percent, od_max, od_max_draw, od_avg, od_avg_win, od_avg_loss]
        td_dat = [td_correct, td_tot, td_perecent, td_max, td_max_draw, td_avg, td_avg_win, td_avg_loss]
        cols = ['Correct_Trades', 'Total_Trades', '%_Correct', 'Max_PnL', 'Max_Drawdown',
                'Avg_Trade', 'Avg_Win', 'Avg_Loss']

        self.trade_metrics = pd.DataFrame([od_dat, td_dat], columns=cols)

    def save_metrics(self, side, param, dfs, sheet_name, stack_row=False):
        self.save_file = f'{self.save_loc}\\predictions_{side}_{param}.xlsx'
        sheet_name = f'{side}_{sheet_name}'

        if os.path.exists(self.save_file):
            # Load the existing workbook
            book = load_workbook(self.save_file)
            if not book.sheetnames:
                book.create_sheet(sheet_name)
                book.active.title = sheet_name

            with pd.ExcelWriter(self.save_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                start_positions = get_excel_sheet_df_positions(dfs, stack_row)
                self.write_metrics_to_excel(writer, dfs, sheet_name, start_positions)

        else:
            # Create a new Excel file
            create_new_excel_file(self.save_file, sheet_name)
            with pd.ExcelWriter(self.save_file, engine='openpyxl') as writer:
                start_positions = get_excel_sheet_df_positions(dfs, stack_row)
                self.write_metrics_to_excel(writer, dfs, sheet_name, start_positions)

    def write_metrics_to_excel(self, writer, dfs, sheet_name, start_positions):
        for df, (startrow, startcol) in zip(dfs, start_positions):
            df.to_excel(writer, sheet_name=sheet_name,
                        startrow=startrow, startcol=startcol)


def get_excel_sheet_df_positions(dfs, stack_row):
    start_positions = [(0, 0)]

    if len(dfs) > 1:
        if stack_row:
            start_row = len(dfs[0])
            for df in dfs[1:]:
                start_row += 2
                start_positions.append((start_row, 0))
                start_row += len(df)
        else:
            start_row = 0
            start_col = len(dfs[0].columns) + 2
            for df in dfs[1:]:
                start_positions.append((start_row, start_col))
                start_row += len(df) + 2

    return start_positions


def create_new_excel_file(file_path, sheet_name):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(file_path)
        wb.close()












